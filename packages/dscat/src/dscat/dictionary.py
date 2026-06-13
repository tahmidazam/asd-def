r"""Read Excel data dictionaries with openpyxl (read-only, values-only).

Cells are normalised to stripped strings with carriage returns removed (the SPARK
dictionaries ship CRLF, which otherwise leaves a stray ``\r`` in header cells).
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

import openpyxl


def _clean(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.replace("\r", "").strip()
    return str(value).strip()


def sheet_names(path: Path) -> list[str]:
    """Return the worksheet names in an Excel workbook, in document order.

    Parameters
    ----------
    path : Path
        Path to the ``.xlsx`` dictionary.

    Returns
    -------
    list of str
        Worksheet names.
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        return list(wb.sheetnames)
    finally:
        wb.close()


def read_sheet(path: Path, sheet: str, max_rows: int | None = None) -> list[list[str]]:
    """Read a worksheet as rows of cleaned string cells.

    Parameters
    ----------
    path : Path
        Path to the ``.xlsx`` dictionary.
    sheet : str
        Worksheet name.
    max_rows : int, optional
        Stop after this many rows; ``None`` reads the whole sheet.

    Returns
    -------
    list of list of str
        One list of stripped string cells per row.
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[sheet]
        rows: list[list[str]] = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if max_rows is not None and i >= max_rows:
                break
            rows.append([_clean(c) for c in row])
        return rows
    finally:
        wb.close()
